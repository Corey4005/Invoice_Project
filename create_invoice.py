#library to import the excel file
import openpyxl
#libraries to create the pdf file and add text to it
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont
#library to get logo related information
from PIL import Image
#convert the font so it is compatible
pdfmetrics.registerFont(TTFont('Arial','Arial.ttf'))

#set the arguments from command line
import sys
workbook = sys.argv[1]
row_start = int(sys.argv[2])
row_end = int(sys.argv[3])

##check for arguments 
if len(sys.argv) != 4:
    print('Usage: create_invoice.py ["excel document name"] ["row to start"] ["row to end"]')
    sys.exit()
else:
    pass

##import the sheet from the excel file
path = r'C:\Users\Corey4005\Documents\Sand Mountain Vertical Farms\Invoices'
wb = openpyxl.load_workbook(path + '\\' + workbook + '.xlsx')

##select the worksheet
sheet = wb['Invoices']

#import company's logo
im = r'C:\Users\Corey4005\Documents\Sand Mountain Vertical Farms\shine_artist-attachments\Logo-Transparency.png'
width, height = 960, 560
ratio = width/height
image_width = 960
image_height = int(image_width / ratio)

#Page information
page_width = 2550
page_height = 3300


#Invoice variables
company_name ='Sand Mountain Micros LLC'
company_address ='872 Fire Tower Rd. Grant, Al. 35747'
margin = 100

#def function
def create_invoice():
    for i in range(row_start, row_end):
        customer = sheet.cell(row = i, column = 2).value
        address = sheet.cell(row = i, column = 5).value
        city = sheet.cell(row = i, column = 6).value
        zip_code = str(sheet.cell(row = i, column = 4).value)
        phone = str(sheet.cell(row = i, column = 13).value)
        email = sheet.cell(row = i, column = 3).value
        invoice_number = sheet.cell(row = i, column = 14).value
        invoice_date = sheet.cell(row = i, column = 15).value
        item1 = sheet.cell(row = i, column = 8).value
        item2 = sheet.cell(row = i, column = 9).value
        item3 = sheet.cell(row = i, column = 10).value
        item4 = sheet.cell(row = i, column = 11).value
        title = 'Memo:' + ' ' + 'Invoice' + ' ' + '#' + str(invoice_number) 

        #Creating a pdf file and setting a naming convention
        c = canvas.Canvas(str(invoice_number) + '_' + str(customer) +'.pdf')
        c.setTitle(title)
        c.setFont('Arial', 40)
        c.drawString(30, 50, title)
        c.setPageSize((page_width, page_height))
        

        #Drawing the image
        c.drawImage(im, x= (page_width - image_width + margin), y= (page_height - image_height - margin), mask='auto')

        #Invoice Start
        c.setFont('Arial',80)
        text = 'INVOICE'
        text_width = stringWidth(text,'Arial',80)
        c.drawString((page_width-text_width)/2, page_height - image_height - 2*margin, text)
        y = page_height - image_height - margin*4
        x = 2*margin
        x2 = x + 550

        #Invoice Info
        c.setFont('Arial', 45)
        c.drawString(x, y, 'Issued by: ')
        c.drawString(x2,y, company_name)
        y -= margin

        c.drawString(x, y, 'Company Address:')
        c.drawString(x2, y, company_address)
        y -= margin
        
        c.drawString(x,y,'Issued to: ')
        c.drawString(x2,y,customer)
        y -= margin

        c.drawString(x,y,'Customer address:')
        c.drawString(x2,y, address + ' ' + city + ' ' + zip_code)
        y -= margin
        
        c.drawString(x,y, 'Invoice date: ')
        c.drawString(x2,y, invoice_date)
        y -= margin
        
        c.drawString(x,y,'Email info:')
        c.drawString(x2, y, email)
        y -= margin

        c.drawString(x,y, 'Phone number:')
        c.drawString(x2, y, phone)
        y -= margin

        #Items Sold
        c.setFont('Arial', 80)
        text2 = 'ITEMS'
        c.drawString((page_width - text_width)/2, y - margin, text2)
        y -= 3*margin
        c.setFont('Arial', 45)
        
        if item1 is not None:
            c.drawString(x, y, 'Item 1')
            c.drawString(x2, y, item1)
            y -= margin

        if item2 is not None:
            c.drawString(x, y, 'Item 2')
            c.drawString(x2, y, item2)
            y -= margin

        if item3 is not None:
            c.drawString(x, y, 'Item 3')
            c.drawString(x2, y, item3)
            y -= margin

        if item4 is not None:
            c.drawString(x, y, 'Item 4')
            c.drawString(x2, y, item4)

        y -= 2*margin
        
        c.drawString(x,y, 'Payment recieved? (circle):')
        c.drawString(x2, y, '     ' + 'Paypal,' + '     ' + 'Venmo,' + '     ' + 'Cash,' + '     ' + 'Check')
        
            

        #Saving the pdf file
        c.save()

def text_create():
    Text = []
    
    for i in range(row_start, row_end):
        phone = str(sheet.cell(row = i, column = 13).value)
        customer = str(sheet.cell(row = i, column = 2).value)
        string = 'Reminder, you will be recieving the following items on'
        at = 'at'
        time = str(sheet.cell(row = i, column = 16).value)
        date = str(sheet.cell(row = i, column = 15).value)
        item1 = str(sheet.cell(row = i, column = 8).value)
        item2 = str(sheet.cell(row = i, column = 9).value)
        item3 = str(sheet.cell(row = i, column = 10).value)
        item4 = str(sheet.cell(row = i, column = 11).value)
        Text.append(phone + ',' + ' ' + customer + ' ' + string + ' ' + at + ' ' + time + ' ' + date + ' ' + item1 + ' ' + item2 + ' ' + item3 + ' ' + item4)
        
    return

        
if __name__ == '__main__':
    create_invoice()
    #text_create()
